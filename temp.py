from openpyxl import Workbook
from openpyxl.formatting import formatting
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.builtins import styles
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule

wb = Workbook()
ws = wb.active

# Create fill
redFill = PatternFill(start_color='EE1111',
               end_color='EE1111',
               fill_type='solid')

# # Add a two-color scale
# # Takes colors in excel 'RRGGBB' style.
# ws.conditional_formatting.add('A1:A10',
#             ColorScaleRule(start_type='min', start_color='AA0000',
#                           end_type='max', end_color='00AA00')
#                           )
#
# # Add a three-color scale
# ws.conditional_formatting.add('B1:B10',
#                ColorScaleRule(start_type='percentile', start_value=10, start_color='AA0000',
#                            mid_type='percentile', mid_value=50, mid_color='0000AA',
#                            end_type='percentile', end_value=90, end_color='00AA00')
#                              )
#
# # Add a conditional formatting based on a cell comparison
# # addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
# # Format if cell is less than 'formula'
# ws.conditional_formatting.add('C2:C10',
#             CellIsRule(operator='lessThan', formula=['C$1'], stopIfTrue=True, fill=redFill))
#
# # Format if cell is between 'formula'
# ws.conditional_formatting.add('D2:D10',
#             CellIsRule(operator='between', formula=['1','5'], stopIfTrue=True, fill=redFill))
#
# # Format using a formula
# ws.conditional_formatting.add('E1:E10',
#             FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=redFill))
#
# # Aside from the 2-color and 3-color scales, format rules take fonts, borders and fills for styling:
# myFont = Font()
# myBorder = Border()
# ws.conditional_formatting.add('E1:E10',
#             FormulaRule(formula=['E1=0'], font=myFont, border=myBorder, fill=redFill))

# Highlight cells that contain particular text by using a special formula
# red_text = Font(color="9C0006")
# red_fill = PatternFill(bgColor="FFC7CE")
# dxf = DifferentialStyle(font=red_text, fill=red_fill)
# rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
#
#
# red_text = Font(color='9C0006')
# red_fill = PatternFill(bgColor='FFC7CE')
# ws.conditional_formatting.add('A1:I44', Rule(type='containsText', operator='containsText', text='gg',fill=red_fill, font=red_text))
#
# # rule.formula = ['NOT(ISERROR(SEARCH("highlight",A)))']
# # ws.conditional_formatting.add('A1:F40', rule)
# blackFill = PatternFill(start_color='000000', end_color='000001', fill_type='solid')
# greyFill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
# colour_dict = {1: blackFill, 2: greyFill}
# for x in range(1, 3):
#     ws.conditional_formatting.add('C2:J12',
#                                      CellIsRule(operator='equal',
#                                                 formula=[x],
#                                                 stopIfTrue=True,
#                                                 fill=colour_dict[x]
#
# red_colour = 'ffc7ce'
# red_colour_font = '9c0103'
#
# red_font = Font(size=14, bold=True, color=red_colour_font)
# red_fill = PatternFill(start_color=red_colour, end_color=red_colour, fill_type='solid')
#
# rule = Rule(type='containsText', text='gg', stopIfTrue=True)
# rule.dxf = DifferentialStyle(font=red_font, border=None, fill=red_fill)
# ws.conditional_formatting.add('A1:E4', rule)
red_color = 'ffc7ce'
red_color_font = '9c0103'

red_font = Font(size=14, bold=True, color=red_color_font)
red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')



ws.conditional_formatting.add('A1:A10',
                              CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))
ws.conditional_formatting.add('B1:B10', CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))

wb.save("test.xlsx")