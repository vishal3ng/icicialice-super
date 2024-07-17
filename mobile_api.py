import datetime
import time
import traceback

from openpyxl.drawing import colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
import openpyxl
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import By
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from selenium.webdriver.support.ui import WebDriverWait as we
from selenium.webdriver.support import expected_conditions as EC

wel = ""
text_ofele = ""
sheet_index = 0

login_memory = {
    1: ['POST', '/content/versionfetchupdate', 'api.icicidirect.com', '11:00:25:298', '666', 'ms', '265', 'B', '/',
        '146', 'B', '200'],
    2: ['POST', '/customer/figpinreg', 'api.icicidirect.com', '11:00:25:311', '1.41', 's', '303', 'B', '/', '74',
        'B', '200'],
    3: ['GET', '/ScripMaster.zip', 'isec-app-masters.s3.dualstack.ap-south-1.amazonaws.com', '11:00:25:343', '5.87',
        's', '0', 'B', '/', '21.34', 'MB', '200'],
    4: ['POST', '/customersetting/verifyfirstfactor/v1', 'api.icicidirect.com', '11:00:39:981', '2.01', 's', '515',
        'B', '/', '299', 'B', '200'],
    5: ['POST', '/customersetting/verifyotpandgensession/v2', 'api.icicidirect.com', '11:00:48:958', '835', 'ms',
        '854', 'B', '/', '906', 'B', '200'],
    6: ['POST', '/customersetting/clientdetails/v1', 'api.icicidirect.com', '11:00:49:867', '177', 'ms', '487', 'B',
        '/', '976', 'B', '200'],
    7: ['POST', '/equity/predefinewtchlst', 'api.icicidirect.com', '11:00:50:75', '515', 'ms', '283', 'B', '/',
        '4.88', 'kB', '200'],
    8: ['POST', '/customersetting/tnc/v1', 'api.icicidirect.com', '11:00:50:76', '135', 'ms', '303', 'B', '/',
        '1.49', 'kB', '200'],
    9: ['POST', '/customer/nomineedeclaration/v1', 'api.icicidirect.com', '11:00:51:408', '495', 'ms', '286', 'B',
        '/', '148', 'B', '200'],
    10: ['POST', '/equity/eqwatchlist', 'api.icicidirect.com', '11:00:51:408', '567', 'ms', '262', 'B', '/', '577',
         'B', '200'],
    11: ['POST', '/equity/getquote', 'api.icicidirect.com', '11:00:51:408', '623', 'ms', '262', 'B', '/', '894',
         'B', '200'],
    12: ['POST', '/equity/getquote', 'api.icicidirect.com', '11:00:51:408', '1.10', 's', '263', 'B', '/', '929',
         'B', '200'],
    13: ['POST', '/equity/getquote', 'api.icicidirect.com', '11:00:51:408', '1.18', 's', '263', 'B', '/', '896',
         'B', '200'],
    14: ['POST', '/customersetting/tnc/v1', 'api.icicidirect.com', '11:00:53:183', '447', 'ms', '306', 'B', '/',
         '55', 'B', '200']}
eqOrderPlace_memory = {
    1: ['POST', '/equity/eqwatchlist', 'api.icicidirect.com', '17:22:10:152', '290', 'ms', '285', 'B', '/', '50', 'B',
        '200'],
    2: ['POST', '/equity/getquote', 'api.icicidirect.com', '17:22:10:152', '1.62', 's', '262', 'B', '/', '892', 'B',
        '200'],
    3: ['POST', '/equity/getquote', 'api.icicidirect.com', '17:22:10:152', '317', 'ms', '263', 'B', '/', '934', 'B',
        '200'],
    4: ['POST', '/equity/getquote', 'api.icicidirect.com', '17:22:20:419', '323', 'ms', '263', 'B', '/', '1.07', 'kB',
        '200'],
    5: ['GET', '/clientapi/ichbiah/stock/overview/YESBANK/', 'ichbiah.trendlyne.com', '17:22:20:481', '399', 'ms', '0',
        'B', '/', '19.08', 'kB', '200'],
    6: ['POST', '/equity/marketdepth', 'api.icicidirect.com', '17:22:20:481', '360', 'ms', '400', 'B', '/', '740', 'B',
        '200'],
    7: ['GET', '/clientapi/ichbiah/stock/fundamental/YESBANK/', 'ichbiah.trendlyne.com', '17:22:20:481', '1.05', 's',
        '0', 'B', '/', '247.72', 'kB', '200'],
    8: ['POST', '/equity/stocklist/v1', 'api.icicidirect.com', '17:22:22:321', '417', 'ms', '305', 'B', '/', '505', 'B',
        '200'],
    9: ['POST', '/equity/order/v2', 'api.icicidirect.com', '17:23:11:412', '517', 'ms', '990', 'B', '/', '268', 'B',
        '200']}

filename = datetime.datetime.now().strftime("%d_%m_%y_%H_%M_%S")


class mobile_api:
    excelName = datetime.datetime.now().strftime("%d-%m-%Y")
    api_button = "(//android.view.View/android.widget.Button[@resource-id=''])[last()]"
    skip_button = "//android.view.View[@content-desc='Skip'] | //android.widget.ImageView[@content-desc='Get Started']"
    accept_alert = "//android.widget.Button[@resource-id='com.android.permissioncontroller:id/permission_allow_button']"
    user_id = "//android.widget.ImageView[@text='User ID']"
    user_password = "//android.widget.ImageView[@text='Password / PIN']"
    login_button = "//android.widget.Button[@content-desc='Login']"
    last_apiOnpage = "//android.view.View[contains(@content-desc,'.com')][last()]"
    first_apiOnpage = "(//android.view.View[contains(@content-desc,'.com')])[2]"
    one_apiOnpage = "//android.view.View[contains(@content-desc,'.com') or contains(@content-desc,'.net')][var]"
    apiOnpage = "//android.view.View[contains(@content-desc,'.com') or contains(@content-desc,'.net')]"
    cancel_biometrics = "//android.widget.Button[@content-desc='Cancel']"
    risk_disk_popup = '//android.widget.Button[@content-desc="Ok, Got it"]'
    api_Show_menu = '//android.widget.Button[@content-desc="Show menu"]'
    api_Delete = '//android.widget.Button[@content-desc="Delete"]'
    api_Yes = '//android.widget.Button[@content-desc="Yes"]'
    api_Back = '//android.widget.Button[@content-desc="Back"]'
    xpathFor_OTP = "//android.widget.TextView[contains(@text,'HLbtq4KhYKi')]"
    Watchlist_tab = "//android.widget.ImageView[contains(@content-desc,'Watchlist')]"
    search_bar = "//android.view.View[@content-desc='Search & add']"
    search_bar_set_value = "//android.widget.EditText[@text='Search Name or Symbol']"
    select_script_base = "//android.view.View[contains(@content-desc,'script') and contains(@content-desc,'exe')]"
    fill_otpxpath = "//android.view.View[@content-desc='OTP']/following-sibling::android.widget.EditText"
    order_actionxpath = "//android.widget.Button[@content-desc='action']"
    clearOtpNotifications = "//android.widget.ImageView[contains(@content-desc,'Clear all notifications')] | //android.widget.Button[contains(@content-desc,'Clear all notifications')] | //android.widget.TextView[contains(@content-desc,'Clear,Button')]"
    placement_exchange = "//android.view.View[contains(@content-desc,'exe')]"
    placement_product = "//android.view.View[contains(@content-desc,'Delivery')]"
    placement_qty = "//android.widget.EditText[1]"
    placement_price = "//android.widget.EditText[last()]"
    placement_Charges_More = "//android.widget.ImageView[@content-desc='Charges & More']"
    placement_available = "//android.view.View[@content-desc='Available Limit']"
    placement_Market_type = "//android.view.View[@content-desc='Market']"
    placement_Limit_type = "//android.view.View[@content-desc='Limit']"
    placement_more = "//android.widget.ImageView[@content-desc='Charges & More']"
    placement_day_validity = "//android.view.View[@content-desc='Day']"
    placement_less = "//android.widget.ImageView[@content-desc='Less']"
    placement_ltp = "//android.view.View[contains(@content-desc,'Price')]"
    orders_tab = "//android.widget.ImageView[contains(@content-desc,'Orders')]"
    orders_tab_pending = "//android.view.View[contains(@content-desc,'PENDING')]"
    # placement_action = "//android.widget.Button[@content-desc='action']"

    desired_capabilities = {
        "platformName": "Android",
        "platformVersion": "14",
        # "deviceName": "e73be1a8",
        "deviceName": "493a3b7d",
        # "deviceId": "192.0.0.2:5555",
        "appPackage": "com.icicisecurities.alice",
        "appActivity": "com.icicisecurities.alice.MainActivity",
        "automationName": "UiAutomator2",
        "ignoreHiddenApiPolicyError": "true",
        "noReset": "false",
        "autoAcceptAlerts": "true"
    }
    script_name = "YESBANK"
    driver = ''
    wait = ''

    def createDriver(self):
        options = UiAutomator2Options().load_capabilities(self.desired_capabilities)
        self.driver = webdriver.Remote("http://localhost:4723/wd/hub", options=options)
        self.driver.implicitly_wait(25)
        self.wait = we(self.driver, 45)

    def swipe_to(self):
        try:
            startX=self.driver.get_window_size()
            print(f"start size {startX}")
            self.driver.swipe(470, 1400, 470, 1000, 400)
        except Exception as e:
            traceback.print_exc()
    def orderplacement(self, order_action, order_type, exchange, qty):
        try:
            self.click_on(self.order_actionxpath.replace("action", order_action))
            self.click_on(self.placement_exchange.replace("exe", exchange))
            self.click_on(self.placement_product)
            self.setValue(self.placement_qty, qty)
            self.click_on(self.placement_Limit_type)
            price = round(float(self.get_text(self.placement_ltp).split()[2].replace("₹", "")), 1)
            print(f"place price {price}")
            self.setValue(self.placement_price, price)
            self.click_on(self.placement_available)
            self.click_on(self.order_actionxpath.replace("action", order_action))
        except Exception as e:
            traceback.print_exc()

    def search_scriptAndSelect(self, script, exchange="NSE"):
        try:
            self.click_on(self.Watchlist_tab)
            self.click_on(self.search_bar)
            self.setValue(self.search_bar_set_value, self.script_name)
            self.click_on(self.select_script_base.replace("script", script).replace("exe", exchange))
        except Exception as e:
            print(str(e))

    def fill_otp(self):
        try:

            self.driver.open_notifications()
            print("notification bar open")
            self.wait.until(EC.visibility_of(self.driver.find_element(By.XPATH, self.xpathFor_OTP)))
            otp_msg = self.get_text(self.xpathFor_OTP, "text")[0:6]
            print(otp_msg)
            self.click_on(self.clearOtpNotifications)
            self.setValue(self.fill_otpxpath, otp_msg)
        except Exception:
            print("unable to set otp")

    def checkDataWithMemory_remark(self, scrapdata, old_data):
        print(f"scrap data from check point : {scrapdata} ")
        print(f"old data from check point : {old_data} ")
        try:
            # Check repeat call api and new api's
            for i in range(1, len(scrapdata) + 1):
                mcqnt = 0
                for ii in range(1, len(old_data) + 1):

                    if scrapdata[i][1] == old_data[ii][1]:
                        mcqnt = mcqnt + 1
                if mcqnt >= 1:
                    scrapdata[i].append("No Change")
                if mcqnt == 0:
                    scrapdata[i].append("New Added")

            for i in range(1, len(scrapdata) + 1):
                mcqnt = 0
                for ii in range(1, len(scrapdata) + 1):
                    if scrapdata[i][1] == scrapdata[ii][1]:
                        mcqnt = mcqnt + 1
                if mcqnt == 1:
                    scrapdata[i].append(" 1 ")
                if mcqnt >= 2:
                    scrapdata[i].append(F"{mcqnt}")

            # if any api not found
            for i in range(1, len(old_data) + 1):
                mcqnt = 0
                newadd = 1
                for ii in range(1, len(scrapdata) + 1):
                    if old_data[i][1] == scrapdata[ii][1]:
                        mcqnt = mcqnt + 1
                if mcqnt == 0:
                    old_data[i].append("Missing")
                    scrapdata[(len(scrapdata) + 1)] = old_data[i]
                    print(f" bfor result {scrapdata}")
        except Exception as e:
            print(f" memory check error {e}  {traceback.print_exc()}")
        return scrapdata

    def clearAndBack_api(self):
        try:
            self.click_on(self.api_button)
            self.click_on(self.api_Show_menu)
            self.click_on(self.api_Delete)
            self.click_on(self.api_Yes)
            self.click_on(self.api_Back)
        except Exception:
            print("unable to clear api ")

    def data_loaderto_dict(self, tcount):
        global new_text
        try:
            data_loader = {}
            runcount = 1
            singledt = ''
            count = len(self.driver.find_elements(By.XPATH, self.apiOnpage))
            print(f"count of api {count}")
            for fir in range(1, count + 1):
                singledt = self.get_text(self.one_apiOnpage.replace("var", str(fir)))
                text = [x for x in singledt.split()]
                data_loader[runcount] = text
                print(f"{runcount}/{tcount}")
                runcount = runcount + 1
            if runcount < tcount:
                while runcount < tcount:
                    print(f"{runcount}/{tcount}")
                    self.swipeFrom_To(self.last_apiOnpage, self.first_apiOnpage)
                    for i in range(1, len(self.driver.find_elements(By.XPATH, self.apiOnpage)) + 1):
                        new_text = self.get_text(self.one_apiOnpage.replace("var", str(i)))
                        if new_text == singledt:
                            ii = i + 1
                            for ch in range(ii, len(self.driver.find_elements(By.XPATH, self.apiOnpage)) + 1):
                                singledt = self.get_text(self.one_apiOnpage.replace("var", str(ch)))
                                text = [x for x in singledt.split()]
                                data_loader[runcount] = text
                                runcount = runcount + 1
                                print(f"{runcount}/{tcount}")
                            break
        except Exception:
            print(f"Fail to load data ")

        # print(f"""----------------------{data_loader}----------------------------""")
        return data_loader

    def data_loaderInExcel(self, apiqnt=1, dtload=None, Sheet_Name="unknown_sheet"):

        try:

            print(f"final data load {dtload}")
            try:
                wb = openpyxl.load_workbook(filename + '.xlsx')
            except Exception as e:
                print(f"new excel file {str(e)}")
                wb = openpyxl.Workbook()
            global sheet_index
            sheet_name2 = Sheet_Name
            wb.create_sheet(sheet_name2, sheet_index)
            sheet_index = sheet_index + 1
            BLUE = 'ADD8E6'
            RED = "00FF0000"
            GREEN = "00FF00"
            YELLOW = "00000080"
            YELLOW = PatternFill(bgColor=RED)
            # style = DifferentialStyle(fill=YELLOW)
            # rule1 = Rule(type="expression", dxf=style)
            # rule1.formula = ['$h="NC"']
            bold_font = Font(bold=True)
            sheet = wb[sheet_name2]
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1'], sheet['F1'], sheet[
                'G1'], sheet[
                'H1'], sheet[
                'I1'] = "request", "Call API.", "Server", "Time to hit API", "Time consume", "Time in", "Response code", "Change ", "Repeat Qty"

            abcd = ['A', 'B', 'C', 'D', 'E', 'F', 'G', "H", "I"]
            for ch2 in abcd:
                sheet[ch2 + '1'].fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
                sheet[ch2 + '1'].font = bold_font
                sheet[ch2 + '1'].border = thin_border
            for key, qlist in dtload.items():
                dtl = list(qlist)
                toset = []
                for ch in range(len(dtl)):
                    if dtl[ch] not in [dtl[dt] for dt in range(6, 11)]:
                        toset.append(dtl[ch])
                sheet.append(toset)
            green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
            red_fill = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
            rule1 = FormulaRule(formula=['ISNUMBER(SEARCH("No Change", C1))'], fill=green_fill)
            rule2 = FormulaRule(formula=['ISNUMBER(SEARCH("New Added", C1))'], fill=red_fill)
            rule3 = FormulaRule(formula=['ISNUMBER(SEARCH("Missing", C1))'], fill=red_fill)
            sheet.conditional_formatting.add('C1:L200', rule1)
            sheet.conditional_formatting.add('C1:L200', rule2)
            sheet.conditional_formatting.add('C1:L200', rule3)
            # ---
            wb.save(filename + '.xlsx')
            print("--------------------- Excel data load done ")
            self.clearAndBack_api()
        except Exception:
            print("unable to load data to excel  cause excel sheet is open ****************")

    # ---------Xpath--------------------------
    def get_text(self, xpath, atribute="content-desc"):
        global text_ofele
        try:
            text_ofele = self.driver.find_element(By.XPATH, xpath).get_attribute(atribute)
        except Exception:
            print(f"unable to get text of element {xpath}")
        # print(f"{text_ofele} text from get text ")
        return text_ofele

    def get_webelemet(self, xpath):
        global wel
        try:
            wel = None
            wel = self.driver.find_element(By.XPATH, xpath)
        except Exception:
            print(f"unable to swipe {xpath}")
        return wel

    def swipeFrom_To(self, fromxpath, toxpath, by_coordinate=None):
        if by_coordinate == None:
            try:
                el1 = self.get_webelemet(fromxpath)
                el2 = self.get_webelemet(toxpath)
                self.driver.scroll(el1, el2)
            except Exception:
                print("unable to swipe ")
        # elif by_coordinate!=None:
        #     try:
        #         actions=
        #         actions.long_press(None,fromxpath,starty).move_to(None,endx,endy).release().perform()
        #     except Exception:
        #         print("unable to swipe ")

    def call_api(self, ):
        pass

    def click_on(self, xpath):
        try:
            self.driver.find_element(By.XPATH, xpath).click()
        except Exception as e:
            print(f"click on {str(e)}")

    def end_activity(self, sheetname, historyData):
        try:
            time.sleep(2)
            total_callApi = int(self.get_text(self.api_button))
            self.click_on(self.api_button)
            time.sleep(2)
            dk = dict(self.data_loaderto_dict(total_callApi))
            data_to_load = self.checkDataWithMemory_remark(dk, historyData)
            self.data_loaderInExcel(dtload=data_to_load, apiqnt=total_callApi, Sheet_Name=sheetname)
            print("done final ")
            time.sleep(1)
            self.driver.quit()
        except Exception:
            print("fail to perform data load activity ")

    def setValue(self, xpath, value):
        try:
            time.sleep(1)
            webel = self.driver.find_element(By.XPATH, xpath).click()
            time.sleep(0.4)
            self.driver.find_element(By.XPATH, xpath).clear()
            time.sleep(0.4)
            self.driver.find_element(By.XPATH, xpath).send_keys(value)
            if self.driver.is_keyboard_shown():
                self.driver.hide_keyboard()
            time.sleep(3)

        except Exception as e:
            print(f"found error msg {str(e)}")

    time.sleep(10)

    def create_dyxpath(self, count, contentForxpath):
        basexpath = f"//android.view.View[contains(@content-desc,'{contentForxpath[0]}')"
        content_to_add = "contains(@content-desc,'var')"
        temp = ""
        for ch in range(1, count):
            updated_Xpath = temp + ' and ' + content_to_add.replace("var", contentForxpath[ch])
            temp = updated_Xpath
        lastxpath = basexpath + updated_Xpath + "]"

    def login_page_api(self, historydata, sheetname=None):
        self.createDriver()
        try:
            print("in mth")
            self.wait.until(EC.alert_is_present())
            self.driver.switch_to.alert.accept()
        except Exception as e:
            print(f"error {e}")
        self.click_on(self.skip_button)
        self.setValue(self.user_id, "170026")
        self.setValue(self.user_password, "170026")
        self.click_on(self.login_button)
        self.fill_otp()
        self.wait.until(EC.visibility_of(self.get_webelemet(self.cancel_biometrics)))
        self.click_on(self.cancel_biometrics)
        self.click_on(self.risk_disk_popup)
        if sheetname != None:
            self.end_activity(sheetname, historyData=historydata)
        if sheetname == None:
            self.clearAndBack_api()

    def place_equityOrder_watchlist(self, historydata, sheetname, scriptName, Exchange, OrderAction, limit, qty):
        self.swipe_to()
        self.login_page_api(historydata=login_memory)
        self.search_scriptAndSelect(scriptName, Exchange)
        self.orderplacement(OrderAction, limit, Exchange, qty)
        self.end_activity(sheetname, historyData=historydata)


login = mobile_api()
# login.login_page_api(login_memory, "login_Page")
login.place_equityOrder_watchlist(eqOrderPlace_memory, "eq_order_placement_watchlist", "YESBANK", "NSE", "Buy", "LIMIT",
                                  "1")
